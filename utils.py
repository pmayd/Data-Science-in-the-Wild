from collections import defaultdict
from typing import List, Union

import openpyxl
import pandas as pd

HEADER_ROWS = range(6, 10)
DIMENSIONS_SHEET_NAME = "Structure"
DIMENSIONS_CELL_RANGE = "B4:E1000"
DIMENSIONS_FIRST_ROW = 3
DIMENSIONS_COL_CAT = 0
DIMENSIONS_COL_CODE = 2
DIMENSIONS_COL_LABEL = 3


def get_header_codes_from_excel(excel_file: str) -> dict:
    """Return Dimension categories, codes and labels from an Eurostat dataset.

    Args:
        excel_file (str): Path to a local excel file.

    Returns:
        dict: A dictionary with categories as key and another dict with code - label as key-value.
    """
    book = openpyxl.load_workbook(excel_file, data_only=True)

    codes = defaultdict(dict)
    sheet = book[DIMENSIONS_SHEET_NAME]
    cells = sheet[DIMENSIONS_CELL_RANGE]

    for row in cells:
        cat = row[DIMENSIONS_COL_CAT].value
        code = row[DIMENSIONS_COL_CODE].value
        label = row[DIMENSIONS_COL_LABEL].value
        if not cat:
            break

        codes[cat][code] = label

    return codes


def print_codes(excel_file: str):
    """Print the Dimension categories, codes and labels from an Eurostat dataset.

    Args:
        excel_file (str): Path to a local excel file.
    """
    codes = get_header_codes_from_excel(excel_file)
    for k, v in codes.items():
        print("Category: ", k)
        print("---------")
        for k, v in v.items():
            print(f"{k}: {v}")
        print()


def get_data_from_excel(
    excel_file: str, headers: Union[tuple, List[tuple]] = None
) -> pd.DataFrame:
    book = openpyxl.load_workbook(excel_file, data_only=True)
    list_of_df = []

    if isinstance(headers, tuple):
        headers = [headers]

    for sheetname in book.sheetnames[2:]:
        sheet = book[sheetname]

        header_names = []
        header_values = []

        for row in HEADER_ROWS:
            name = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=3).value
            
            if name is None:
                break
            
            name = name.split("[")[1].strip("]")
            value = value.split("[")[1].strip("]")
            header_names.append(name)
            header_values.append(value)
        
        header_names = tuple(header_names)
        header_values = tuple(header_values)

        # if we search for a specific header, skip sheet if no header matches the current sheet header
        # if we call the function with no headers, we take all sheets
        if headers:
            if not any(header_values == header for header in headers):
                continue

        # drop column GEO(L)/TIME
        # rename GEO to geo
        # melt dataframe from wide to long with columns year and value
        # add four columns for unit, hazard, waste and nace_r2
        # change columns year and geo to type category
        # set geo as index

        # find header row
        header_row = 0
        for row in range(1, 21):
            value = sheet.cell(row=row, column=1).value
            if value is not None and value.startswith("TIME"):
                header_row = row
                break

        nrows = header_row
        for row in range(header_row, 101):
            if sheet.cell(row=row, column=1).value is None:
                nrows = row - header_row - 1
                break
                    
        df_sheet = pd.read_excel(
            excel_file,
            sheet_name=sheetname,
            header=header_row-1,
            nrows=nrows,
            na_values=":",
        )
        df_sheet = df_sheet.rename(
            columns={
                df_sheet.columns[0]: df_sheet.iloc[0, 0],
                df_sheet.columns[1]: df_sheet.iloc[0, 1],
            }
        )
        df_sheet = df_sheet.iloc[1:]
        df_sheet = df_sheet.drop(columns="GEO (Labels)")
        df_sheet = df_sheet.rename(columns={"GEO (Codes)": "geo"})
        df_sheet = df_sheet.melt(id_vars="geo", var_name="year", value_name="value")
        df_sheet = df_sheet.assign(
            **{x.lower(): y for x, y in zip(header_names, header_values) if x}
        )

        list_of_df.append(df_sheet)

    df = pd.concat(list_of_df)
    df.year = df.year.astype(int)
    df[df.select_dtypes("object").columns] = df.select_dtypes("object").astype(
        "category"
    )
    df = df.set_index("geo")

    return df
